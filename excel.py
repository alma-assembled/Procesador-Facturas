import json
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from Models.opModel import opModel
from Models.CuentasCobrar import ModelFacturasCobrar , CuentasPorCobrar 
from Models.CuentrasPagar import ModelFacturasPagar
from Models.produccion import Produccion
from Models.Insumos import Insumos


class ExcelGenerator:
    
    def crear_excel_desde_mysql(self, folio):
        self.model_op = opModel()
        self.modelCobrar = ModelFacturasCobrar()
        self.modelPagar = ModelFacturasPagar()
        self.produccion = Produccion()
        self.insumos = Insumos()
        HOJA_RESUMEN = "Resumen"
        HOJA_COBRAR = "Facturas por Cobrar"
        HOJA_PAGAR = "Facturas por pagar"
        HOJA_INSUMOS = "Insumos Almacen"
        HOJA_SALARIOS = "Salarios Produccion"

        datos_cobrar, jsonCobrar = self.modelCobrar.facturasCobraroDetalles(folio)
        datos_pagar, jsonPagar =  self.modelPagar.facturasPagarDetalles(folio)
        datos_insumos, jsonInsumos =  self.insumos.insumosOp(folio)
        opDatos=self.model_op.ConsultaOpByFolio(folio)
        libro = openpyxl.Workbook()

    
        # HOJA_COBRAR
        hoja_cobrar = libro.create_sheet(HOJA_COBRAR)
        encabezados =  list(json.loads(jsonCobrar)[0].keys())
        hoja_cobrar.append(encabezados)

        
        columnas_cobrar = ['A', 'B', 'C', 'D', 'E', 'F']
        self.formatoExcel(columnas_cobrar,hoja_cobrar)

        for fila_json in json.loads(jsonCobrar):
            fila_excel = [fila_json[encabezado] for encabezado in encabezados]
            hoja_cobrar.append(fila_excel)

        #HOJA_PAGAR
        hoja_pagar = libro.create_sheet(HOJA_PAGAR)
        encabezados =  list(json.loads(jsonPagar)[0].keys())
        hoja_pagar.append(encabezados)


        columnas_pagar = ['A', 'B', 'C', 'D', 'E', 'F']
        self.formatoExcel(columnas_pagar,hoja_pagar)

        for fila_json in json.loads(jsonPagar):
            fila_excel = [fila_json[encabezado] for encabezado in encabezados]
            hoja_pagar.append(fila_excel)

        #HOJA_ISUMOS-ALMANCEN
        
        hoja_insumos = libro.create_sheet(HOJA_INSUMOS)
        encabezados =  list(json.loads(jsonInsumos)[0].keys())
        hoja_insumos.append(encabezados)


        columnas_insumos = ['A', 'B', 'C', 'D', 'E', 'F']
        self.formatoExcel(columnas_insumos,hoja_insumos)

        for fila_json in json.loads(jsonInsumos):
            fila_excel = [fila_json[encabezado] for encabezado in encabezados]
            hoja_insumos.append(fila_excel)
        
        #HOJA_SALARIOS
        ajustes, maquila, produccion  = self.produccion.totalProduccion("714")
        hoja_salarios = libro.create_sheet(HOJA_SALARIOS)
        total_maquila = ajustes + maquila + produccion
        hoja_salarios['C4'] = 'Pago Ajustes' 
        hoja_salarios['C5'] = 'Pago Maquiladores'
        hoja_salarios['C6'] = 'Pago Produccion'
        hoja_salarios['C7'] = 'Total Maquila'

        hoja_salarios['D4'] = ajustes
        hoja_salarios['D5'] = maquila
        hoja_salarios['D6'] = produccion
        hoja_salarios['D7'] = total_maquila

        #HOJA_RESUMEN
        hoja_resumen = libro.create_sheet(HOJA_RESUMEN)

        hoja_resumen['C4']='Folio'
        hoja_resumen['C5']='Cliente'
        hoja_resumen['C6']='Nombre'
        hoja_resumen['C7']='Fecha liberacion'
        hoja_resumen['C8']='Fecha entega'

        hoja_resumen['D4']=opDatos[1]
        hoja_resumen['D5']=self.model_op.ConsultaClienteOp(folio)[0]
        hoja_resumen['D6']=opDatos[2]
        hoja_resumen['D7']=opDatos[4]
        hoja_resumen['D8']=opDatos[3]


        hoja_resumen['G4']= 'Productos'
        productos = self.model_op.ConsultaProductos(folio)
        n=4
        for producto in  productos :
            hoja_resumen['H'+str(n)] = producto[0]
            n += 1

        hoja_resumen['K4'] = 'Facturas Cobrar' 
        hoja_resumen['K5'] = 'Facturas Pagar'
        hoja_resumen['K6'] = 'Insumos Almacen'
        hoja_resumen['K7'] = 'Salarios Maquila'
        hoja_resumen['k8'] = 'Utilidad'

        hoja_resumen['L4'] = 16529
        hoja_resumen['L5'] = 2997
        hoja_resumen['L6'] = 20
        hoja_resumen['L7'] = total_maquila
        hoja_resumen['L8'] = '=L4-SUMA(L5:L7)'


        # Elimina la hoja inicial predeterminada
        libro.remove(libro.active)

        # Guarda el libro de Excel
        libro.save("ResumenOp.xlsx")

    def formatoExcel(self, columnas, hoja):
        # Aplicar estilo a los encabezados
        for cell in hoja[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(bottom=Side(style='thin'))


        for col in columnas:
            hoja.column_dimensions[col].width = 25

        for row in hoja.iter_rows(min_row=2, max_row=hoja.max_row, min_col=4, max_col=6):
            for cell in row:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')



if __name__ == "__main__":
    excel_generator = ExcelGenerator()
    
    excel_generator.crear_excel_desde_mysql("714")